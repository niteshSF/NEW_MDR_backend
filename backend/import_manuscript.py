"""
Bulk Import Script for Acharya Database
Reads data from the Excel template and imports into MySQL database
"""

import openpyxl
from sqlalchemy import create_engine, text, null
from sqlalchemy.orm import sessionmaker
import sys
from datetime import datetime
from typing import Dict, List, Tuple, Optional


class MDRBulkImporter:
    def __init__(self, db_url: str, excel_file: str):
        """
        Initialize the bulk importer

        Args:
            db_url: Database connection URL (e.g., 'mysql+pymysql://user:pass@localhost/acharya')
            excel_file: Path to the Excel template file
        """
        self.db_url = db_url
        self.excel_file = excel_file
        self.engine = create_engine(db_url)
        Session = sessionmaker(bind=self.engine)
        self.session = Session()

        # Cache for lookups
        self.book_cache: Dict[str, int] = {}
        self.manuscript_cache: Dict[str, int] = {}
        self.branch_cache: Dict[str, int] = {}
        self.category_cache: Dict[str, int] = {}
        self.discipline_cache: Dict[str, int] = {}
        self.language_cache: Dict[str, int] = {}
        self.script_cache: Dict[str, int] = {}
        self.subject_cache: Dict[str, int] = {}
        self.tag_cache: Dict[str, int] = {}
        self.type_cache: Dict[str, int] = {}

        # Statistics
        self.stats = {
            "manuscripts": {"inserted": 0, "errors": 0},
            "tags": {"inserted": 0, "errors": 0},
            "manuscript_tag": {"inserted": 0, "errors": 0},
            "books": {"inserted": 0, "errors": 0},
        }

        self.errors: List[Tuple[str, int, str]] = []  # (sheet, row, error_message)

    def load_lookups(self):
        """Load existing data into cache"""
        print("Loading existing data into cache...")

        # Load books
        result = self.session.execute(
            text("SELECT published_title, id, publisher_name, editor_name FROM book")
        )

        for row in result:
            book_key = f"{row.published_title}|{row.publisher_name}|{row.editor_name}"
            self.book_cache[book_key] = row.id
        print(f"Loaded {len(self.book_cache)} books")

        # Load manuscripts
        result = self.session.execute(text("SELECT * FROM manuscript"))
        for row in result:
            manuscript_key = f"{row.name}|{row.accession_number}|{row.subject_id}"
        #     self.manuscript_cache[manuscript_key] = row.id
        # print(f"Loaded {len(self.manuscript_cache)} manuscripts")

        # Load branch
        result = self.session.execute(text("SELECT id, name FROM branch"))
        for row in result:
            self.branch_cache[row.name] = row.id
        print(f"Loaded {len(self.branch_cache)} branches")

        # Load categories
        result = self.session.execute(text("SELECT id, name FROM category"))
        for row in result:
            self.category_cache[row.name] = row.id
        print(f"Loaded {len(self.category_cache)} categories")

        # Load discipline
        result = self.session.execute(text("SELECT id, name FROM discipline"))
        for row in result:
            self.discipline_cache[row.name] = row.id
        print(f"Loaded {len(self.discipline_cache)} disciplines")

        # Load language
        result = self.session.execute(text("SELECT id, name FROM language"))
        for row in result:
            self.language_cache[row.name] = row.id
        print(f"Loaded {len(self.language_cache)} languages")

        # Load script
        result = self.session.execute(text("SELECT id, name FROM script"))
        for row in result:
            self.script_cache[row.name] = row.id
        print(f"Loaded {len(self.script_cache)} scripts")

        # Load subject
        result = self.session.execute(text("SELECT id, name FROM subject"))
        for row in result:
            self.subject_cache[row.name] = row.id
        print(f"Loaded {len(self.subject_cache)} subjects")

        # Load tag
        result = self.session.execute(text("SELECT id, name FROM tag"))
        for row in result:
            self.tag_cache[row.name] = row.id
        print(f"Loaded {len(self.tag_cache)} tags")

        # Load type
        result = self.session.execute(text("SELECT id, name FROM type"))
        for row in result:
            self.type_cache[row.name] = row.id
        print(f"Loaded {len(self.type_cache)} types")

    def parse_branch_reference(self, reference: str) -> Optional[int]:
        """
        Parse branch reference and return branch_id
        Format: branch_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.branch_cache:
            return self.branch_cache[reference]

        return None

    def parse_category_reference(self, reference: str) -> Optional[int]:
        """
        Parse category reference and return category_id
        Format: category_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.category_cache:
            return self.category_cache[reference]

        return None

    def parse_discipline_reference(self, reference: str) -> Optional[int]:
        """
        Parse discipline reference and return discipline_id
        Format: discipline_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.discipline_cache:
            return self.discipline_cache[reference]

        return None

    def parse_language_reference(self, reference: str) -> Optional[int]:
        """
        Parse language reference and return language_id
        Format: language_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.language_cache:
            return self.language_cache[reference]

        return None

    def parse_script_reference(self, reference: str) -> Optional[int]:
        """
        Parse script reference and return script_id
        Format: script_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.script_cache:
            return self.script_cache[reference]

        return None

    def parse_subject_reference(self, reference: str) -> Optional[int]:
        """
        Parse subject reference and return subject_id
        Format: subject_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.subject_cache:
            return self.subject_cache[reference]

        return None

    def parse_tag_reference(self, reference: str) -> Optional[int]:
        """
        Parse tag reference and return tag_id
        Format: tag_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.tag_cache:
            return self.tag_cache[reference]

        return None

    def parse_type_reference(self, reference: str) -> Optional[int]:
        """
        Parse type reference and return type_id
        Format: type_name
        """
        if not reference or reference.strip() == "":
            return None

        if reference in self.type_cache:
            return self.type_cache[reference]

        return None

    def import_tags(self, ws) -> bool:
        """Import tags from each sheet"""
        print("\nImporting Tags...")
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            # Skip empty rows
            if not row[0].value:
                continue

            try:
                # Split the string into a list using a comma as the delimiter
                tag_value = str(row[10].value) if row[10].value else ""
                tag_list = tag_value.split(",")

                created_by = 1  # Assuming admin user ID is 1
                created_at = datetime.now()
                updated_by = 1
                updated_at = datetime.now()
                active = True

                # Loop through the list of items
                for item in tag_list:
                    tag_id = self.parse_tag_reference(item.strip())
                    tag_name = item.strip()
                    if tag_id:
                        print(
                            f"Row {row_idx}: Tag '{item.strip()}' found, skipping insertion"
                        )
                        continue

                    if not tag_id:
                        # Insert manuscripttaglink
                        query = text(
                            """
                            INSERT INTO tag (name, short_name, created_by, created_at, updated_by, updated_at, active)
                            VALUES (:name, :short_name, :created_by, :created_at, :updated_by, :updated_at, :active)
                        """
                        )
                        result = self.session.execute(
                            query,
                            {
                                "name": item.strip(),
                                "short_name": "",
                                "created_by": created_by,
                                "created_at": created_at,
                                "updated_by": updated_by,
                                "updated_at": updated_at,
                                "active": active,
                            },
                        )

                        tag_id = result.lastrowid
                        self.tag_cache[tag_name] = tag_id
                        self.stats["tags"]["inserted"] += 1
                        print(
                            f"Row {row_idx}: Inserted tag '{tag_name}' (ID: {tag_id})"
                        )

            except Exception as e:
                self.stats["tags"]["errors"] += 1
                error_msg = f"Error: {str(e)}"
                self.errors.append(("tags", row_idx, error_msg))
                print(f"Row {row_idx}: {error_msg}")

        self.session.commit()
        return True

    def import_book(self, ws) -> bool:
        """Import books from each sheet"""
        print("\nImporting Books...")
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            # Skip empty rows
            if not row[2].value:
                continue

            try:
                is_published = row[24].value == "Yes" or False
                if not is_published:
                    continue

                translator_name = row[26].value
                publisher_name = row[27].value
                editor_name = row[28].value
                publication_year = row[29].value
                publication_place = row[30].value
                no_of_pages = row[31].value
                archive_link = row[32].value
                beginning_line = row[33].value
                ending_line = row[34].value
                colophon = row[35].value

                book_name = (
                    str(row[25].value).strip()
                    if row[25].value
                    else str(row[2].value).strip()
                )
                if not book_name:
                    continue

                book_key = f"{book_name}|{publisher_name}|{editor_name}"
                if book_key in self.book_cache:
                    print(f"Row {row_idx}: Book {book_key} already exists, skipping")
                    continue

                created_by = 1  # Assuming admin user ID is 1
                created_at = datetime.now()
                updated_by = 1
                updated_at = datetime.now()
                active = True

                # Insert book
                query = text(
                    """
                    INSERT INTO book (published_title, created_by, created_at, updated_by, updated_at, active,
                             translator_name, publisher_name, editor_name, publication_year, publication_place, no_of_pages, archive_link, beginning_line, ending_line, colophon)
                    VALUES (:published_title, :created_by, :created_at, :updated_by, :updated_at, :active,
                            :translator_name, :publisher_name, :editor_name, :publication_year, :publication_place, :no_of_pages, :archive_link, :beginning_line, :ending_line, :colophon)
                """
                )
                result = self.session.execute(
                    query,
                    {
                        "published_title": book_name,
                        "created_by": created_by,
                        "created_at": created_at,
                        "updated_by": updated_by,
                        "updated_at": updated_at,
                        "active": active,
                        "translator_name": translator_name,
                        "publisher_name": publisher_name,
                        "editor_name": editor_name,
                        "publication_year": publication_year,
                        "publication_place": publication_place,
                        "no_of_pages": no_of_pages,
                        "archive_link": archive_link,
                        "beginning_line": beginning_line,
                        "ending_line": ending_line,
                        "colophon": colophon,
                    },
                )

                book_id = result.lastrowid
                self.stats["books"]["inserted"] += 1
                self.book_cache[book_key] = book_id
                print(f"Row {row_idx}: Inserted book '{book_name}' (ID: {book_id})")

            except Exception as e:
                error_msg = f"Error: {str(e)}"
                self.errors.append(("books", row_idx, error_msg))
                print(f"Row {row_idx}: {error_msg}")

        self.session.commit()
        return True

    def import_manuscripts(self, ws) -> bool:
        """Import manuscripts from each sheet"""
        print("\nImporting Manuscripts...")

        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            # Skip empty rows
            if not row[2].value:
                continue

            try:
                accession_num = row[1].value
                name = row[2].value
                diacritical_name = row[3].value
                indic_name = row[4].value

                # Get language_id
                language_id = self.parse_language_reference(row[5].value)
                if not language_id:
                    raise ValueError(f"Language '{row[5].value}' not found")

                # Get script_id
                script_id = self.parse_script_reference(row[6].value)
                if not script_id:
                    raise ValueError(f"Script '{row[6].value}' not found")

                # Get subject_id
                subject_id = self.parse_subject_reference(row[7].value)
                if not subject_id:
                    raise ValueError(f"Subject '{row[7].value}' not found")

                # Get category_id
                category_id = self.parse_category_reference(row[8].value)
                if not category_id:
                    raise ValueError(f"Category '{row[8].value}' not found")

                # Get type_id
                type_id = self.parse_type_reference(row[9].value)
                if not type_id:
                    raise ValueError(f"Type '{row[9].value}' not found")

                # Get book_id
                book_name = (
                    str(row[25].value).strip()
                    if row[25].value
                    else str(row[2].value).strip()
                )
                publisher_name = row[27].value
                editor_name = row[28].value
                book_id = self.book_cache.get(
                    f"{book_name}|{publisher_name}|{editor_name}", None
                )
                if not book_id:
                    print(
                        f"Row {row_idx}: Book reference '{book_name}' not found, skipping book association"
                    )

                # Get tag_id
                """ tag_id = self.parse_tag_reference(row[10].value)
                if row[10].value and not tag_id:
                    raise ValueError(f"Tag '{row[10].value}' not found") """

                # Check if manuscript already exists
                manuscript_key = f"{name}|{accession_num}|{subject_id}"
                # if manuscript_key in self.manuscript_cache:
                #     print(
                #         f"Row {row_idx}: Manuscript {manuscript_key} already exists, skipping"
                #     )
                #     continue

                summary = row[11].value
                toc = row[12].value
                date_of_composition = row[18].value
                source = row[19].value
                pg_in_source = row[20].value
                ms_code = row[21].value
                is_complete = row[22].value == "Complete" or False
                no_of_folios = row[23].value
                is_published = row[24].value == "Yes" or False

                created_by = 1  # Assuming admin user ID is 1
                created_at = datetime.now()
                updated_by = 1
                updated_at = datetime.now()
                active = True

                # Insert manuscript
                query = text(
                    """
                    INSERT INTO manuscript (accession_number, name, diacritical_name, indic_name, language_id, script_id, subject_id, category_id, type_id,
                              summary, toc, manuscript_code, is_complete, no_of_folios, date_of_composition, source, pg_in_source, published, created_by, 
                             book_id, created_at, updated_by, updated_at, active)
                    VALUES (:accession_num, :name, :diacritical_name, :indic_name, :language_id, :script_id, :subject_id, :category_id, :type_id, 
                             :summary, :toc, :ms_code, :is_complete, :no_of_folios,:date_of_composition,:source,:pg_in_source,:is_published,:created_by,
                             :book_id,:created_at,:updated_by,:updated_at,:active)
                """
                )

                result = self.session.execute(
                    query,
                    {
                        "accession_num": accession_num,
                        "name": name,
                        "diacritical_name": diacritical_name,
                        "indic_name": indic_name,
                        "language_id": language_id,
                        "script_id": script_id,
                        "subject_id": subject_id,
                        "category_id": category_id,
                        "type_id": type_id,
                        "summary": summary,
                        "toc": toc,
                        "date_of_composition": date_of_composition,
                        "source": source,
                        "pg_in_source": pg_in_source,
                        "ms_code": ms_code,
                        "is_complete": is_complete,
                        "no_of_folios": no_of_folios,
                        "is_published": is_published,
                        "book_id": book_id,
                        "created_by": created_by,
                        "created_at": created_at,
                        "updated_by": updated_by,
                        "updated_at": updated_at,
                        "active": active,
                    },
                )

                manuscript_id = result.lastrowid
                self.manuscript_cache[manuscript_key] = manuscript_id
                self.stats["manuscripts"]["inserted"] += 1
                print(
                    f"Row {row_idx}: Inserted manuscript {manuscript_key} (ID: {manuscript_id})"
                )

                tag_value = str(row[10].value) if row[10].value else ""

                # Split the string into a list using a comma as the delimiter
                tag_list = tag_value.split(",")

                # Loop through the list of items
                for item in tag_list:
                    item = item.strip()
                    if not item:
                        continue
                    tag_id = self.parse_tag_reference(item)
                    if not tag_id:
                        print(f"Row {row_idx}: Tag '{item}' not found, skipping")
                        continue

                    if tag_id:
                        # Insert manuscripttaglink
                        query = text(
                            """
                            INSERT INTO manuscripttaglink (manuscript_id, tag_id)
                            VALUES (:manuscript_id, :tag_id)
                        """
                        )
                        self.session.execute(
                            query, {"manuscript_id": manuscript_id, "tag_id": tag_id}
                        )

                        self.stats["manuscript_tag"]["inserted"] += 1
                        print(
                            f"Row {row_idx}: Inserted tag link '{item.strip()}' (ID: {manuscript_id})"
                        )

                # Insert ms_additional_info
                subject_contribution = row[13].value
                work_uniqueness = row[14].value
                author_name = row[15].value
                author_diacritical = row[16].value
                author_indic = row[17].value
                notes = row[36].value

                query = text(
                    """
                    INSERT INTO msadditionalinfo (manuscript_id, subject_contribution, work_uniqueness, author_name, author_diacritical_name, 
                             author_indic_name, created_by, created_at, updated_by, updated_at, active)
                    VALUES (:manuscript_id, :subject_contribution, :work_uniqueness, :author_name, :author_diacritical, 
                             :author_indic, :created_by, :created_at, :updated_by, :updated_at, :active)
                """
                )

                result = self.session.execute(
                    query,
                    {
                        "manuscript_id": manuscript_id,
                        "subject_contribution": subject_contribution,
                        "work_uniqueness": work_uniqueness,
                        "author_name": author_name,
                        "author_diacritical": author_diacritical,
                        "author_indic": author_indic,
                        "created_by": created_by,
                        "created_at": created_at,
                        "updated_by": updated_by,
                        "updated_at": updated_at,
                        "active": active,
                    },
                )

                # self.stats['ms_additional_info']['inserted'] += 1
                print(
                    f"Row {row_idx}: Inserted MS Additional Info {manuscript_key} (ID: {manuscript_id})"
                )

            except Exception as e:
                self.stats["manuscripts"]["errors"] += 1
                error_msg = f"Error: {str(e)}"
                self.errors.append(("Manuscripts", row_idx, error_msg))
                print(f"Row {row_idx}: {error_msg}")

        self.session.commit()
        return True

    def run_import(self):
        """Run the complete import manuscript process"""
        print("=" * 60)
        print("MDR Manuscript Bulk Import Tool")
        print("=" * 60)
        print(f"Excel file: {self.excel_file}")
        print(f"Database: {self.db_url.split('@')[-1]}")  # Hide credentials
        print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)

        try:
            # Load workbook
            print("\nLoading Excel file...")
            wb = openpyxl.load_workbook(self.excel_file)
            print(f"Found {len(wb.sheetnames)} sheets")

            # Load caches
            self.load_lookups()

            # Import in order (respecting foreign key dependencies)
            for sheet in wb.sheetnames:
                if not "Status" in sheet:
                    self.import_book(wb[sheet])
                    self.import_tags(wb[sheet])
                    self.import_manuscripts(wb[sheet])

            # Print summary
            print("\n" + "=" * 60)
            print("IMPORT SUMMARY")
            print("=" * 60)

            total_inserted = 0
            total_errors = 0

            for table, counts in self.stats.items():
                inserted = counts["inserted"]
                errors = counts["errors"]
                total_inserted += inserted
                total_errors += errors

                if inserted > 0 or errors > 0:
                    status = "✓" if errors == 0 else "⚠"
                    print(
                        f"{status} {table.capitalize():20} - Inserted: {inserted:4}, Errors: {errors:4}"
                    )

            print("-" * 60)
            print(
                f"  TOTAL:                   Inserted: {total_inserted:4}, Errors: {total_errors:4}"
            )
            print("=" * 60)

            # Print errors if any
            if self.errors:
                print("\nERRORS ENCOUNTERED:")
                print("-" * 60)
                for sheet, row, error in self.errors:
                    print(f"  {sheet} - Row {row}: {error}")

            print(f"\nCompleted at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        except Exception as e:
            print(f"\nFATAL ERROR: {str(e)}")
            self.session.rollback()
            raise
        finally:
            self.session.close()


def main():
    """Main entry point"""
    if len(sys.argv) < 2:
        print("Usage: python import_manuscript.py <excel_file>")
        print("\nExample:")
        print("python import_manuscript.py import_manuscript_template.xlsx")
        sys.exit(1)

    db_url = "mysql+pymysql://root:root@localhost/new_mdr"  # sys.argv[1]
    excel_file = sys.argv[1]

    importer = MDRBulkImporter(db_url, excel_file)
    importer.run_import()


if __name__ == "__main__":
    main()
